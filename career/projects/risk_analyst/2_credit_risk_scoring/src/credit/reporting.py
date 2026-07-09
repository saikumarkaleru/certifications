"""Reporting layer: Excel workbook + matplotlib charts.

Everything an analyst hands over: a formatted workbook (Inputs / Altman /
Merton / Portfolio) and two charts (Altman Z bar chart by zone; Merton PD vs
Altman Z scatter).
"""

from __future__ import annotations

import os

import matplotlib
matplotlib.use("Agg")            # save charts without a display
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ZONE_COLORS = {"Safe": "seagreen", "Grey": "goldenrod",
               "Distress": "firebrick", "N/A": "grey"}

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


# =========================================================================
# Excel
# =========================================================================
def _style_sheet(worksheet, money_cols=None, pct_cols=None, num_cols=None):
    """Bold header row, freeze it, auto-width, and apply number formats."""
    money_cols = money_cols or []
    pct_cols = pct_cols or []
    num_cols = num_cols or []

    for cell in worksheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    worksheet.freeze_panes = "A2"

    # header names -> column index
    headers = {cell.value: cell.column for cell in worksheet[1]}

    def _fmt(col_names, fmt):
        for name in col_names:
            if name in headers:
                letter = get_column_letter(headers[name])
                for cell in worksheet[letter][1:]:
                    cell.number_format = fmt

    _fmt(money_cols, "#,##0")
    _fmt(pct_cols, "0.00%")
    _fmt(num_cols, "0.000")

    # auto width
    for column_cells in worksheet.columns:
        length = max((len(str(c.value)) for c in column_cells if c.value is not None),
                     default=8)
        letter = get_column_letter(column_cells[0].column)
        worksheet.column_dimensions[letter].width = min(max(length + 2, 10), 24)


def write_excel(path: str, inputs: pd.DataFrame, altman: pd.DataFrame,
                merton: pd.DataFrame, combined: pd.DataFrame,
                el_table: pd.DataFrame, summary: dict) -> str:
    """Write the four-sheet workbook and return the path."""
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        inputs.to_excel(writer, sheet_name="Inputs")
        altman.to_excel(writer, sheet_name="Altman")
        merton.to_excel(writer, sheet_name="Merton")

        # Portfolio sheet: combined ranking on top, EL below, totals at bottom.
        combined.to_excel(writer, sheet_name="Portfolio", startrow=0)
        el_start = len(combined) + 3
        el_table.to_excel(writer, sheet_name="Portfolio", startrow=el_start)
        totals = pd.DataFrame({
            "Metric": ["Total EAD", "Total Expected Loss", "EL % of exposure"],
            "Value": [summary["total_ead"], summary["total_el"],
                      summary["el_pct"]],
        })
        totals.to_excel(writer, sheet_name="Portfolio",
                        startrow=el_start + len(el_table) + 3, index=False)

        book = writer.book
        _style_sheet(book["Inputs"], money_cols=list(inputs.columns[:-1]),
                     num_cols=["equity_vol"])
        _style_sheet(book["Altman"],
                     num_cols=["X1", "X2", "X3", "X4", "X5", "Z"])
        _style_sheet(book["Merton"],
                     money_cols=["default_point", "asset_value"],
                     pct_cols=["asset_vol", "PD"], num_cols=["DD"])

    return path


# =========================================================================
# Charts
# =========================================================================
def chart_altman_bar(altman: pd.DataFrame, path: str) -> str:
    """Bar chart of Z by company, coloured by zone, with threshold lines."""
    from .altman import DISTRESS_THRESHOLD, SAFE_THRESHOLD

    data = altman.sort_values("Z", ascending=False)
    colors = [ZONE_COLORS.get(z, "grey") for z in data["Zone"]]

    fig, ax = plt.subplots(figsize=(11, 6))
    ax.bar(data.index, data["Z"], color=colors, edgecolor="black")
    ax.axhline(SAFE_THRESHOLD, color="green", linestyle="--", linewidth=1.4,
               label=f"Safe threshold ({SAFE_THRESHOLD})")
    ax.axhline(DISTRESS_THRESHOLD, color="red", linestyle="--", linewidth=1.4,
               label=f"Distress threshold ({DISTRESS_THRESHOLD})")

    for i, (name, row) in enumerate(data.iterrows()):
        ax.text(i, row["Z"] + 0.08, f"{row['Z']:.2f}", ha="center",
                va="bottom", fontsize=9, fontweight="bold")

    ax.set_title("Altman Z-Score by Company (coloured by risk zone)")
    ax.set_ylabel("Z-Score")
    ax.set_xlabel("Company")
    ax.legend()
    fig.tight_layout()
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    fig.savefig(path, dpi=120)
    plt.close(fig)
    return path


def chart_pd_vs_z(altman: pd.DataFrame, merton: pd.DataFrame,
                  path: str) -> str:
    """Scatter of Merton PD (y) vs Altman Z (x), labelled/coloured by zone."""
    from .altman import DISTRESS_THRESHOLD, SAFE_THRESHOLD

    joined = altman.join(merton[["PD"]])
    fig, ax = plt.subplots(figsize=(9, 6.5))
    for name, row in joined.iterrows():
        color = ZONE_COLORS.get(row["Zone"], "grey")
        ax.scatter(row["Z"], row["PD"] * 100, color=color, s=90,
                   edgecolor="black", zorder=3)
        ax.annotate(name, (row["Z"], row["PD"] * 100),
                    textcoords="offset points", xytext=(6, 4), fontsize=9)

    ax.axvline(SAFE_THRESHOLD, color="green", linestyle="--", linewidth=1.2)
    ax.axvline(DISTRESS_THRESHOLD, color="red", linestyle="--", linewidth=1.2)
    ax.set_title("Structural vs Accounting: Merton PD vs Altman Z")
    ax.set_xlabel("Altman Z-Score (higher = safer)")
    ax.set_ylabel("Merton 1yr Probability of Default (%)")

    # simple zone legend
    handles = [plt.Line2D([0], [0], marker="o", color="w", label=z,
                          markerfacecolor=c, markersize=9, markeredgecolor="black")
               for z, c in ZONE_COLORS.items() if z != "N/A"]
    ax.legend(handles=handles, title="Altman zone")
    fig.tight_layout()
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    fig.savefig(path, dpi=120)
    plt.close(fig)
    return path
