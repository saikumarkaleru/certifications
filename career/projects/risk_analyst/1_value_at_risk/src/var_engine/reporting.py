"""Reporting layer: an Excel workbook and three matplotlib charts."""

from __future__ import annotations

import os

import matplotlib
matplotlib.use("Agg")  # headless backend; no display required
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autosize(ws):
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = width + 3


def write_excel(path, returns, var_summary, component_df, backtest, source, notional):
    """Write the four-sheet workbook: Returns, VaR_Summary, Component, Backtest."""
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb = Workbook()

    # --- VaR_Summary ---
    ws = wb.active
    ws.title = "VaR_Summary"
    ws.append(["Method", "Confidence", "VaR %", "VaR $", "ES %", "ES $"])
    for row in var_summary:
        ws.append(row)
    _style_header(ws)
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 2).number_format = "0%"
        ws.cell(r, 3).number_format = "0.0000"
        ws.cell(r, 4).number_format = "$#,##0"
        ws.cell(r, 5).number_format = "0.0000"
        ws.cell(r, 6).number_format = "$#,##0"
    ws.cell(ws.max_row + 2, 1, f"Data source: {source}")
    ws.cell(ws.max_row + 1, 1, f"Portfolio notional: ${notional:,.0f}")
    _autosize(ws)

    # --- Component ---
    ws = wb.create_sheet("Component")
    ws.append(["Ticker", "Weight", "Marginal VaR", "Component VaR", "% Contribution"])
    for _, r in component_df.iterrows():
        ws.append([r["ticker"], r["weight"], r["marginal_var"],
                   r["component_var"], r["pct_contribution"]])
    _style_header(ws)
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 2).number_format = "0.0%"
        ws.cell(r, 3).number_format = "0.000000"
        ws.cell(r, 4).number_format = "0.000000"
        ws.cell(r, 5).number_format = "0.0%"
    _autosize(ws)

    # --- Backtest ---
    ws = wb.create_sheet("Backtest")
    ws.append(["Metric", "Value"])
    rows = [
        ("Confidence", f"{backtest.conf:.0%}"),
        ("Observations", backtest.n_obs),
        ("Exceptions", backtest.n_exceptions),
        ("Expected exceptions", round(backtest.expected, 2)),
        ("Kupiec LR_uc", round(backtest.lr_uc, 4)),
        ("Kupiec p-value", round(backtest.p_uc, 4)),
        ("Christoffersen LR_ind", round(backtest.lr_ind, 4)),
        ("Christoffersen p-value", round(backtest.p_ind, 4)),
        ("Combined LR_cc", round(backtest.lr_cc, 4)),
        ("Combined p-value", round(backtest.p_cc, 4)),
        ("Verdict (5%)", backtest.verdict()),
    ]
    for row in rows:
        ws.append(list(row))
    _style_header(ws)
    _autosize(ws)

    # --- Returns (tail of the series to keep file light) ---
    ws = wb.create_sheet("Returns")
    ws.append(["Date"] + list(returns.columns))
    tail = returns.tail(250)
    for idx, r in tail.iterrows():
        ws.append([pd.Timestamp(idx).strftime("%Y-%m-%d")] + [float(v) for v in r.values])
    _style_header(ws)
    for r in range(2, ws.max_row + 1):
        for c in range(2, ws.max_column + 1):
            ws.cell(r, c).number_format = "0.0000"
    _autosize(ws)

    wb.save(path)
    return path


def chart_distribution(port_returns, var95, var99, path):
    """Histogram of portfolio returns with 95% and 99% VaR lines."""
    fig, ax = plt.subplots(figsize=(9, 5))
    ax.hist(np.asarray(port_returns), bins=60, color="#4C72B0", alpha=0.8, edgecolor="white")
    ax.axvline(-var95, color="#DD8452", lw=2, ls="--", label=f"95% VaR = {var95:.2%}")
    ax.axvline(-var99, color="#C44E52", lw=2, ls="--", label=f"99% VaR = {var99:.2%}")
    ax.set_title("Portfolio Daily Return Distribution with VaR")
    ax.set_xlabel("Daily return")
    ax.set_ylabel("Frequency")
    ax.legend()
    fig.tight_layout()
    fig.savefig(path, dpi=110)
    plt.close(fig)
    return path


def chart_rolling_var(rolling_var, path):
    """Line chart of the rolling historical VaR."""
    fig, ax = plt.subplots(figsize=(9, 5))
    ax.plot(rolling_var.index, rolling_var.values * 100.0, color="#55A868", lw=1.5)
    ax.set_title("Rolling 250-Day Historical VaR (95%)")
    ax.set_xlabel("Date")
    ax.set_ylabel("VaR (% of portfolio)")
    ax.grid(True, alpha=0.3)
    fig.autofmt_xdate()
    fig.tight_layout()
    fig.savefig(path, dpi=110)
    plt.close(fig)
    return path


def chart_risk_contribution(component_df, path):
    """Bar chart of component VaR by asset."""
    fig, ax = plt.subplots(figsize=(9, 5))
    ax.bar(component_df["ticker"], component_df["component_var"] * 100.0, color="#8172B3")
    ax.set_title("Risk Contribution by Asset (Component VaR, 95%)")
    ax.set_xlabel("Ticker")
    ax.set_ylabel("Component VaR (% of portfolio)")
    for i, v in enumerate(component_df["component_var"] * 100.0):
        ax.text(i, v, f"{v:.2f}", ha="center", va="bottom", fontsize=8)
    fig.tight_layout()
    fig.savefig(path, dpi=110)
    plt.close(fig)
    return path
