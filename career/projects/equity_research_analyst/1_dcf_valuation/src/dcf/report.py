"""
report.py -- write the formatted Excel workbook and the two charts.

Sheets:
  Assumptions  -- every driver + the WACC build.
  DCF          -- the per-year FCFF projection, discounting and the EV->equity
                  bridge.
  Scenarios    -- bull / base / bear value/share.
  Sensitivity  -- 2-way WACC x terminal-growth grid.
  ReverseDCF   -- the growth the market is implying vs our base case.

Charts:
  fcf_bridge.png    -- projected vs discounted FCFF (with terminal value).
  football_field.png -- valuation range (scenarios + reverse) vs market price.
"""

from __future__ import annotations

import os

import numpy as np
import matplotlib
matplotlib.use("Agg")  # headless backend: no display needed
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ---- shared styling ----
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=13, color="1F4E78")
BOLD = Font(bold=True)
PCT = "0.0%"
CUR = "#,##0"
CUR2 = "#,##0.00"


def _style_header(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _num(cell, fmt):
    cell.number_format = fmt


def write_workbook(path, company, wacc_info, fcff_rows, base, scenarios,
                   sens, reverse, model_assumptions):
    wb = Workbook()

    # ================= Sheet 1: Assumptions =================
    ws = wb.active
    ws.title = "Assumptions"
    ws["A1"] = f"DCF Assumptions -- {company['name']} ({company['ticker']})"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Data source: {company.get('source')}  |  as of {company.get('as_of')}"

    rows = [
        ("Current market price ($)", company["price"], CUR2),
        ("Shares outstanding", company["shares"], CUR),
        ("Net debt ($)", company.get("net_debt"), CUR),
        ("Beta", wacc_info["beta"], "0.00"),
        ("Risk-free rate (10Y ^TNX)", wacc_info["risk_free_rate"], PCT),
        ("Equity risk premium", wacc_info["erp"], PCT),
        ("Cost of equity (CAPM)", wacc_info["cost_of_equity"], PCT),
        ("Cost of debt (pre-tax)", wacc_info["cost_of_debt_pre_tax"], PCT),
        ("Tax rate", wacc_info["tax_rate"], PCT),
        ("Cost of debt (after-tax)", wacc_info["cost_of_debt_after_tax"], PCT),
        ("Weight of equity", wacc_info["weight_equity"], PCT),
        ("Weight of debt", wacc_info["weight_debt"], PCT),
        ("WACC (discount rate)", wacc_info["wacc"], PCT),
        ("Base FCFF ($)", model_assumptions["base_fcff"], CUR),
        ("Year-1 growth", model_assumptions["start_growth"], PCT),
        ("Terminal growth", model_assumptions["terminal_growth"], PCT),
        ("Exit EV/EBITDA multiple", model_assumptions["exit_multiple"], "0.0x"),
        ("Forecast years", model_assumptions["years"], "0"),
    ]
    ws.cell(row=4, column=1, value="Driver").font = BOLD
    ws.cell(row=4, column=2, value="Value").font = BOLD
    _style_header(ws, 4, 2)
    for i, (label, val, fmt) in enumerate(rows, start=5):
        ws.cell(row=i, column=1, value=label)
        c = ws.cell(row=i, column=2, value=(float(val) if val is not None else None))
        _num(c, fmt if fmt != "0.0x" else '0.0"x"')
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    # ================= Sheet 2: DCF =================
    ws2 = wb.create_sheet("DCF")
    ws2["A1"] = "Projected & Discounted FCFF"
    ws2["A1"].font = TITLE_FONT
    headers = ["Year", "Growth", "Projected FCFF", "Discount Factor", "PV of FCFF"]
    for j, h in enumerate(headers, start=1):
        ws2.cell(row=3, column=j, value=h)
    _style_header(ws2, 3, len(headers))
    r = base["result"] if "result" in base else base
    for i in range(len(r["years"])):
        ws2.cell(row=4 + i, column=1, value=int(r["years"][i]))
        _num(ws2.cell(row=4 + i, column=2, value=float(r["growth_path"][i])), PCT)
        _num(ws2.cell(row=4 + i, column=3, value=float(r["projected_fcff"][i])), CUR)
        _num(ws2.cell(row=4 + i, column=4, value=float(r["discount_factors"][i])), "0.0000")
        _num(ws2.cell(row=4 + i, column=5, value=float(r["pv_fcff"][i])), CUR)

    bridge_start = 4 + len(r["years"]) + 2
    ws2.cell(row=bridge_start, column=1, value="ENTERPRISE -> EQUITY BRIDGE").font = BOLD
    bridge = [
        ("PV of explicit FCFF", r["pv_explicit"], CUR),
        ("Terminal value (Gordon)", r["tv_gordon"], CUR),
        ("Terminal value (exit mult.)", r["tv_exit"], CUR),
        ("Blended terminal value", r["terminal_value"], CUR),
        ("PV of terminal value", r["pv_terminal"], CUR),
        ("Terminal % of EV", r["terminal_pct_of_ev"], PCT),
        ("ENTERPRISE VALUE", r["enterprise_value"], CUR),
        ("Less: net debt", -(company.get("net_debt") or 0.0), CUR),
        ("EQUITY VALUE", r["equity_value"], CUR),
        ("Shares outstanding", company["shares"], CUR),
        ("Intrinsic value / share", r["value_per_share"], CUR2),
        ("Current market price", company["price"], CUR2),
        ("Upside / (downside)", r["value_per_share"] / company["price"] - 1, PCT),
    ]
    for k, (label, val, fmt) in enumerate(bridge, start=bridge_start + 1):
        ws2.cell(row=k, column=1, value=label)
        if val is not None:
            _num(ws2.cell(row=k, column=2, value=float(val)), fmt)
    ws2.column_dimensions["A"].width = 28
    for col in ("B", "C", "D", "E"):
        ws2.column_dimensions[col].width = 18

    # ================= Sheet 3: Scenarios =================
    ws3 = wb.create_sheet("Scenarios")
    ws3["A1"] = "Scenario Analysis (bull / base / bear)"
    ws3["A1"].font = TITLE_FONT
    sh = ["Scenario", "Year-1 Growth", "Terminal g", "WACC",
          "Value / Share", "Upside vs Price"]
    for j, h in enumerate(sh, start=1):
        ws3.cell(row=3, column=j, value=h)
    _style_header(ws3, 3, len(sh))
    for i, name in enumerate(["Bull", "Base", "Bear"], start=4):
        s = scenarios[name]
        ws3.cell(row=i, column=1, value=name)
        _num(ws3.cell(row=i, column=2, value=s["start_growth"]), PCT)
        _num(ws3.cell(row=i, column=3, value=s["terminal_growth"]), PCT)
        _num(ws3.cell(row=i, column=4, value=s["wacc"]), PCT)
        _num(ws3.cell(row=i, column=5, value=s["value_per_share"]), CUR2)
        _num(ws3.cell(row=i, column=6,
                      value=s["value_per_share"] / company["price"] - 1), PCT)
    ws3.column_dimensions["A"].width = 12
    for col in ("B", "C", "D", "E", "F"):
        ws3.column_dimensions[col].width = 16

    # ================= Sheet 4: Sensitivity =================
    ws4 = wb.create_sheet("Sensitivity")
    ws4["A1"] = "Intrinsic Value / Share  (rows = WACC, cols = terminal growth)"
    ws4["A1"].font = TITLE_FONT
    wacc_vals, term_vals, matrix = sens
    ws4.cell(row=3, column=1, value="WACC \\ g").font = BOLD
    for j, g in enumerate(term_vals, start=2):
        _num(ws4.cell(row=3, column=j, value=g), PCT)
    _style_header(ws4, 3, len(term_vals) + 1)
    for i, w in enumerate(wacc_vals, start=4):
        _num(ws4.cell(row=i, column=1, value=w), PCT)
        ws4.cell(row=i, column=1).font = BOLD
        for j, g in enumerate(term_vals, start=2):
            v = matrix[i - 4][j - 2]
            cell = ws4.cell(row=i, column=j,
                            value=(None if np.isnan(v) else float(v)))
            _num(cell, CUR2)
    ws4.column_dimensions["A"].width = 12

    # ================= Sheet 5: ReverseDCF =================
    ws5 = wb.create_sheet("ReverseDCF")
    ws5["A1"] = "Reverse DCF -- what growth is the market pricing in?"
    ws5["A1"].font = TITLE_FONT
    ig = reverse.get("implied_growth")
    rows5 = [
        ("Current market price ($)", company["price"], CUR2),
        ("Our base-case year-1 growth", model_assumptions["start_growth"], PCT),
        ("Market-implied year-1 growth", ig, PCT),
        ("Interpretation", reverse.get("note"), None),
    ]
    for i, (label, val, fmt) in enumerate(rows5, start=3):
        ws5.cell(row=i, column=1, value=label)
        if fmt is None:
            ws5.cell(row=i, column=2, value=val)
        elif val is not None:
            _num(ws5.cell(row=i, column=2, value=float(val)), fmt)
    if ig is not None:
        verdict = ("Market is MORE optimistic than us" if ig >
                   model_assumptions["start_growth"] else
                   "Market is LESS optimistic than us")
        ws5.cell(row=8, column=1, value="Read-through").font = BOLD
        ws5.cell(row=8, column=2, value=verdict)
    ws5.column_dimensions["A"].width = 32
    ws5.column_dimensions["B"].width = 34

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Charts
# ---------------------------------------------------------------------------
def chart_fcf_bridge(path, base, company):
    """Bar chart: projected FCFF vs its present value, per year + terminal PV."""
    r = base["result"] if "result" in base else base
    years = [f"Y{int(y)}" for y in r["years"]]
    proj = r["projected_fcff"] / 1e9
    pv = r["pv_fcff"] / 1e9
    x = np.arange(len(years))

    fig, ax = plt.subplots(figsize=(9, 5))
    ax.bar(x - 0.2, proj, width=0.4, label="Projected FCFF", color="#9DC3E6")
    ax.bar(x + 0.2, pv, width=0.4, label="PV of FCFF", color="#1F4E78")
    # append the discounted terminal value as a separate bar
    ax.bar(len(years) + 0.0, r["pv_terminal"] / 1e9, width=0.4,
           label="PV terminal value", color="#C55A11")
    labels = years + ["Terminal"]
    ax.set_xticks(list(x) + [len(years)])
    ax.set_xticklabels(labels)
    ax.set_ylabel("$ billions")
    ax.set_title(f"FCFF Bridge -- {company['name']} ({company['ticker']})")
    ax.legend()
    plt.tight_layout()
    fig.savefig(path, dpi=120)
    plt.close(fig)
    return path


def chart_football_field(path, scenarios, reverse, company, base_growth,
                         reverse_value=None):
    """Horizontal range chart: scenario values + market price line."""
    names = ["Bear", "Base", "Bull"]
    vals = [scenarios[n]["value_per_share"] for n in names]
    y = np.arange(len(names))

    fig, ax = plt.subplots(figsize=(9, 4.8))
    colors = ["#C55A11", "#1F4E78", "#2E7D32"]
    ax.barh(y, vals, color=colors, height=0.55)
    ax.set_yticks(y)
    ax.set_yticklabels(names)
    ax.invert_yaxis()
    for i, v in enumerate(vals):
        ax.text(v, i, f" ${v:,.0f}", va="center", fontsize=10)
    price = company["price"]
    ax.axvline(price, color="black", linestyle="--", linewidth=1.5)
    ax.text(price, -0.6, f" market ${price:,.0f}", color="black", fontsize=9)
    ax.set_xlabel("Intrinsic value per share ($)")
    ax.set_title(f"DCF Football Field -- {company['name']} ({company['ticker']})")
    plt.tight_layout()
    fig.savefig(path, dpi=120)
    plt.close(fig)
    return path
