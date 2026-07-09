"""
report.py -- write the formatted Excel workbook and the football-field chart.

Sheets:
  Comps      -- raw peer financials (market cap, EV, revenue, EBITDA, etc.).
  Multiples  -- per-company multiples + the peer median/mean/quartile summary.
  Implied    -- implied price per method + the football-field range vs price.
  Screen     -- z-score screen + OLS regression (rich/cheap vs fundamentals).

Chart:
  football_field.png -- implied prices per method + a market-price line.
"""

from __future__ import annotations

import os

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=13, color="1F4E78")
BOLD = Font(bold=True)


def _style_header(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _to_native(v):
    if v is None:
        return None
    if isinstance(v, (np.integer,)):
        return int(v)
    if isinstance(v, (np.floating, float)):
        return None if (isinstance(v, float) and v != v) else float(v)
    return v


def _write_df(ws, df, start_row=3, index_label=None, round_to=2):
    cols = list(df.columns)
    offset = 1 if index_label else 0
    if index_label:
        ws.cell(row=start_row, column=1, value=index_label)
    for j, c in enumerate(cols, start=1 + offset):
        ws.cell(row=start_row, column=j, value=str(c))
    _style_header(ws, start_row, len(cols) + offset)
    for i, (idx, row) in enumerate(df.iterrows(), start=start_row + 1):
        if index_label:
            ws.cell(row=i, column=1, value=str(idx)).font = BOLD
        for j, c in enumerate(cols, start=1 + offset):
            v = _to_native(row[c])
            if isinstance(v, float) and round_to is not None:
                v = round(v, round_to)
            ws.cell(row=i, column=j, value=v)
    return start_row + 1 + len(df)


def write_workbook(path, universe, mult_df, summary, implied_df, football,
                   zscreen, reg_df, reg_coef, reg_r2):
    wb = Workbook()
    target = universe["target"]

    # ---- Sheet 1: Comps (raw peer financials) ----
    ws = wb.active
    ws.title = "Comps"
    ws["A1"] = f"Peer Financials ($) -- target: {target}"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Data source: {universe.get('source')} | as of {universe.get('as_of')}"
    comp_rows = {}
    for tk, c in universe["companies"].items():
        mc = c.get("market_cap") or (c["price"] * c["shares"])
        ev = mc + (c.get("total_debt") or 0) - (c.get("cash") or 0)
        comp_rows[tk] = {
            "name": c.get("name"), "price": c.get("price"),
            "market_cap": mc, "EV": ev, "revenue": c.get("revenue"),
            "ebitda": c.get("ebitda"), "net_income": c.get("net_income"),
            "book_equity": c.get("book_equity"),
            "rev_growth": c.get("rev_growth"),
            "ebitda_margin": c.get("ebitda_margin"),
        }
    comp_df = pd.DataFrame.from_dict(comp_rows, orient="index")
    comp_df.index.name = "ticker"
    _write_df(ws, comp_df.reset_index(), start_row=4, round_to=4)
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 26

    # ---- Sheet 2: Multiples ----
    ws2 = wb.create_sheet("Multiples")
    ws2["A1"] = "Trading Multiples (by company)"
    ws2["A1"].font = TITLE_FONT
    show = mult_df[["name", "P/E", "EV/EBITDA", "EV/Revenue", "P/B", "PEG"]]
    end = _write_df(ws2, show.reset_index(), start_row=3, round_to=2)
    ws2.cell(row=end + 1, column=1, value="Peer summary (peers only)").font = BOLD
    _write_df(ws2, summary, start_row=end + 2, index_label="Multiple", round_to=2)
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 26

    # ---- Sheet 3: Implied ----
    ws3 = wb.create_sheet("Implied")
    ws3["A1"] = f"Implied Valuation of {target} (peer median multiples)"
    ws3["A1"].font = TITLE_FONT
    _write_df(ws3, implied_df, start_row=3, round_to=2)
    fb = 3 + len(implied_df) + 2
    ws3.cell(row=fb, column=1, value="FOOTBALL FIELD").font = BOLD
    fb_rows = [
        ("Low implied price", football["low"]),
        ("Median implied price", football["median"]),
        ("High implied price", football["high"]),
        ("Current market price", football["current_price"]),
        ("Upside to median", football["upside_to_median"]),
        ("Verdict", football["verdict"]),
    ]
    for k, (label, val) in enumerate(fb_rows, start=fb + 1):
        ws3.cell(row=k, column=1, value=label)
        cell = ws3.cell(row=k, column=2, value=_to_native(val))
        if label == "Upside to median":
            cell.number_format = "0.0%"
        elif isinstance(val, float):
            cell.number_format = "#,##0.00"
    ws3.column_dimensions["A"].width = 24
    ws3.column_dimensions["B"].width = 18

    # ---- Sheet 4: Screen ----
    ws4 = wb.create_sheet("Screen")
    ws4["A1"] = "Rich / Cheap Screen"
    ws4["A1"].font = TITLE_FONT
    ws4.cell(row=3, column=1, value="(a) Z-score of raw multiples across the "
             "universe").font = BOLD
    end = _write_df(ws4, zscreen.reset_index(), start_row=4, round_to=2)
    ws4.cell(row=end + 1, column=1,
             value="(b) OLS: EV/EBITDA ~ revenue growth + EBITDA margin").font = BOLD
    if not reg_df.empty:
        end2 = _write_df(ws4, reg_df.reset_index(), start_row=end + 2, round_to=2)
        ws4.cell(row=end2 + 1, column=1, value="Regression coefficients").font = BOLD
        r = end2 + 2
        ws4.cell(row=r, column=1, value="intercept")
        ws4.cell(row=r, column=2, value=round(reg_coef.get("intercept", 0), 2))
        ws4.cell(row=r + 1, column=1, value="beta_growth")
        ws4.cell(row=r + 1, column=2,
                 value=round(reg_coef.get("rev_growth", 0), 2))
        ws4.cell(row=r + 2, column=1, value="beta_margin")
        ws4.cell(row=r + 2, column=2,
                 value=round(reg_coef.get("ebitda_margin", 0), 2))
        ws4.cell(row=r + 3, column=1, value="R-squared")
        ws4.cell(row=r + 3, column=2, value=round(reg_r2, 3))
    ws4.column_dimensions["A"].width = 24

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    return path


def chart_football_field(path, implied_df, football, universe):
    """Horizontal bars: implied price per method + a market-price line."""
    methods = implied_df["Method"].tolist()
    vals = implied_df["Implied Price/Share"].tolist()
    y = np.arange(len(methods))

    fig, ax = plt.subplots(figsize=(9, 4.8))
    ax.barh(y, vals, color="#1F4E78", height=0.55)
    ax.set_yticks(y)
    ax.set_yticklabels(methods)
    ax.invert_yaxis()
    for i, v in enumerate(vals):
        ax.text(v, i, f" ${v:,.0f}", va="center", fontsize=10)
    price = football["current_price"]
    ax.axvline(price, color="crimson", linestyle="--", linewidth=1.6)
    ax.text(price, -0.6, f" market ${price:,.0f}", color="crimson", fontsize=9)
    # shade the low-high range
    ax.axvspan(football["low"], football["high"], color="#9DC3E6", alpha=0.20)
    ax.set_xlabel("Implied price per share ($)")
    tgt = universe["companies"][universe["target"]]
    ax.set_title(f"Comps Football Field -- {tgt.get('name')} "
                 f"({universe['target']})")
    plt.tight_layout()
    fig.savefig(path, dpi=120)
    plt.close(fig)
    return path
