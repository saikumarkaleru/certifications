"""Reporting layer: a formatted Excel workbook and three matplotlib charts."""

from __future__ import annotations

import os

import matplotlib
matplotlib.use("Agg")  # headless backend; no display required
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
FLAG_FILL = PatternFill("solid", fgColor="FCE4D6")  # light amber for exceptions


def _style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autosize(ws):
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 3, 55)


def _write_df(ws, df: pd.DataFrame):
    ws.append(list(df.columns))
    for _, r in df.iterrows():
        ws.append([_cell(v) for v in r.values])
    _style_header(ws)
    _autosize(ws)


def _cell(v):
    if isinstance(v, pd.Timestamp):
        return v.strftime("%Y-%m-%d")
    if pd.isna(v):
        return ""
    return v


def write_excel(path, nav, valuation, trade_breaks, cash_breaks, scored, alerts, fund):
    """Write the six-sheet workbook."""
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb = Workbook()

    # --- NAV_Summary ---
    ws = wb.active
    ws.title = "NAV_Summary"
    ws.append(["Item", "Value"])
    rows = [
        ("Fund", fund["fund_name"]),
        ("Valuation date", fund["valuation_date"].strftime("%Y-%m-%d")),
        ("Holdings market value", nav.holdings_mv),
        ("Cash", nav.cash),
        ("Accrued income", nav.accrued_income),
        ("Gross Asset Value (GAV)", nav.gross_asset_value),
        ("Management fee accrued (TER)", -nav.management_fee_accrued),
        ("Other accrued expenses", -nav.other_accrued_expenses),
        ("Total liabilities", -nav.total_liabilities),
        ("Net Asset Value (NAV)", nav.nav),
        ("Units outstanding", nav.units_outstanding),
        ("NAV per unit", round(nav.nav_per_unit, 6)),
        ("Prior NAV per unit", round(nav.prior_nav_per_unit, 6)),
        ("Day-over-day move", nav.nav_move_pct),
        ("Move flagged (>=2%)", "YES" if nav.move_flagged else "no"),
    ]
    for r in rows:
        ws.append(list(r))
    _style_header(ws)
    for i in range(2, ws.max_row + 1):
        label = ws.cell(i, 1).value
        val = ws.cell(i, 2)
        if isinstance(val.value, (int, float)):
            if label == "Day-over-day move":
                val.number_format = "0.00%"
            elif label in ("Units outstanding",):
                val.number_format = "#,##0"
            elif label == "NAV per unit" or label == "Prior NAV per unit":
                val.number_format = "0.000000"
            else:
                val.number_format = "$#,##0.00"
    _autosize(ws)

    # --- Pricing (per-position valuation) ---
    val_out = valuation[["security_id", "ticker", "asset_class", "quantity", "price",
                         "price_source", "price_age_days", "market_value"]].copy()
    ws = wb.create_sheet("Pricing")
    _write_df(ws, val_out)
    for i in range(2, ws.max_row + 1):
        ws.cell(i, 5).number_format = "0.0000"
        ws.cell(i, 8).number_format = "$#,##0.00"
        if ws.cell(i, 6).value != "GOOD":
            for c in range(1, ws.max_column + 1):
                ws.cell(i, c).fill = FLAG_FILL

    # --- Trade_Breaks & Cash_Breaks ---
    for title, q in (("Trade_Breaks", trade_breaks), ("Cash_Breaks", cash_breaks)):
        ws = wb.create_sheet(title)
        if q.empty:
            ws.append(["(no breaks)"])
        else:
            _write_df(ws, q)
            for i in range(2, ws.max_row + 1):
                if ws.cell(i, 2).value == "High" or ws.cell(i, 3).value == "High":
                    pass
                ws.cell(i, 4).number_format = "$#,##0.00"
                ws.cell(i, 5).number_format = "$#,##0.00"
                ws.cell(i, 6).number_format = "$#,##0.00"

    # --- KYC_Risk ---
    ws = wb.create_sheet("KYC_Risk")
    _write_df(ws, scored)
    tier_col = list(scored.columns).index("risk_tier") + 1
    for i in range(2, ws.max_row + 1):
        if ws.cell(i, tier_col).value == "High":
            for c in range(1, ws.max_column + 1):
                ws.cell(i, c).fill = FLAG_FILL

    # --- AML_Alerts ---
    ws = wb.create_sheet("AML_Alerts")
    if alerts.empty:
        ws.append(["(no alerts)"])
    else:
        _write_df(ws, alerts)

    wb.save(path)
    return path


# --- Charts ----------------------------------------------------------------
def chart_nav_composition(nav, path):
    """Waterfall-style bar of the components bridging GAV to NAV."""
    comp = nav.composition()
    labels = list(comp.keys())
    values = list(comp.values())
    colors = ["#4C72B0" if v >= 0 else "#C44E52" for v in values]
    fig, ax = plt.subplots(figsize=(9, 5))
    ax.bar(labels, values, color=colors)
    ax.axhline(0, color="black", lw=0.8)
    ax.set_title(f"NAV Composition — NAV ${nav.nav:,.0f} ({nav.nav_per_unit:.4f}/unit)")
    ax.set_ylabel("USD")
    for i, v in enumerate(values):
        ax.text(i, v, f"${v/1e6:,.2f}M", ha="center",
                va="bottom" if v >= 0 else "top", fontsize=8)
    plt.xticks(rotation=20, ha="right")
    fig.tight_layout()
    fig.savefig(path, dpi=110)
    plt.close(fig)
    return path


def chart_break_aging(aging_df, path):
    """Bar chart of reconciliation breaks by aging bucket."""
    fig, ax = plt.subplots(figsize=(9, 5))
    ax.bar(aging_df["aging_bucket"], aging_df["count"], color="#DD8452")
    ax.set_title("Reconciliation Break Aging (trade + cash exceptions)")
    ax.set_xlabel("Age bucket")
    ax.set_ylabel("Number of breaks")
    for i, v in enumerate(aging_df["count"]):
        ax.text(i, v, str(int(v)), ha="center", va="bottom", fontsize=9)
    fig.tight_layout()
    fig.savefig(path, dpi=110)
    plt.close(fig)
    return path


def chart_risk_tiers(tier_df, path):
    """Bar chart of the KYC customer risk-tier distribution."""
    colors = {"Low": "#55A868", "Medium": "#DD8452", "High": "#C44E52"}
    fig, ax = plt.subplots(figsize=(9, 5))
    ax.bar(tier_df["risk_tier"], tier_df["count"],
           color=[colors[t] for t in tier_df["risk_tier"]])
    ax.set_title("KYC Customer Risk-Tier Distribution")
    ax.set_xlabel("Risk tier")
    ax.set_ylabel("Number of customers")
    for i, v in enumerate(tier_df["count"]):
        ax.text(i, v, str(int(v)), ha="center", va="bottom", fontsize=9)
    fig.tight_layout()
    fig.savefig(path, dpi=110)
    plt.close(fig)
    return path
